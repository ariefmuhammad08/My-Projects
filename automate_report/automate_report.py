import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList

input_file= 'input_data/supermarket_sales.xlsx'
output_file= 'output_data/report_penjualan2019.xlsx'
wh_url= 'https://discord.com/api/webhooks/1024913462406676520/fng2Zpr-XDeKJPauYLe74bGrcFuA5Wiwd2NL9qyPfsHVPMXkkC7CGVf_DXMgmTK0UuZb'

df = pd.read_excel(input_file)

## print sample data
# print(f'Dataframe colums:{df.columns}')
# print(f'Sample dataset{df.head()}')

## penjualan total per gender & product line
# print(df[['Gender','Product line','Total']].head())

## filter kolom gender hanya male
# print(df['Gender'].loc[df['Gender']=='Male'])

## membuat tabel baru dari yang sudah ada
df=df.pivot_table(index='Gender',
                columns='Product line',
                values='Total',
                aggfunc='sum').round()

## save tabel ke excel
print('save dataframe to excel...')

df.to_excel(output_file,
            sheet_name='Report',
            startrow=4)

print('save dataframe done...')

## membuat grafik
wb= load_workbook(output_file)
wb.active= wb['Report']

min_column= wb.active.min_column
max_column= wb.active.max_column
min_row= wb.active.min_row
max_row= wb.active.max_row

## Mengetahui min/max kolom dan baris
# print("""
#         min_column: {0}
#         max_column: {1}
#         min_row: {2}
#         max_row: {3}
#         """.format(min_column,max_column,min_row,max_row))

barchart = BarChart()

data= Reference(wb.active,
                 min_col= min_column+1,
                 max_col= max_column,
                 min_row=min_row,
                 max_row=max_row)

categories= Reference(wb.active,
                       min_col= min_column,
                       max_col= min_column,
                       min_row= min_row+1,
                       max_row= max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

wb.active.add_chart(barchart,'B12')
barchart.title= 'Sales Berdasarkan Produk'
barchart.style= 2

wb.save(output_file)

## menambahakan total dari penjualan
import string
alphabet= list(string.ascii_uppercase)
alphabet_excel= alphabet[:max_column]

for i in alphabet_excel:
    if i !='A':
        wb.active[f'{i}{max_row+1}']= f'=SUM({i}{min_row+1}:{i}{max_row})'
        wb.active[f'{i}{max_row+1}'].style= 'Currency'

wb.active[f'{alphabet_excel[0]}{max_row+1}']='Total'

wb.save(output_file)

## kirim ke discord
import discord
from discord import SyncWebhook

webhook= SyncWebhook.from_url(wh_url)

with open(file=output_file, mode='rb')as file:
    test_file= discord.File(file)

webhook.send('testing report', username='test', file=test_file)

## input dan output filenya ada di 
## https://github.com/ariefmuhammad08/digitalskola